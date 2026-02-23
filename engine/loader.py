"""
Carga y parseo de archivos Excel de inventario.
"""
import openpyxl
from io import BytesIO
from .styles import COL_MAP_NO_HEADERS, safe_int, safe_float, safe_str
from .detector import detect_headers, detect_line_and_sizes


def load_inventory(file_obj, progress_callback=None):
    """Carga inventario desde Excel (BytesIO o ruta).

    Args:
        file_obj: BytesIO (de Streamlit upload) o str (ruta de archivo)
        progress_callback: función(pct, msg) para reportar progreso

    Returns:
        dict con keys: records, has_headers, linea, tallas, total_rows
    """
    if progress_callback:
        progress_callback(0.05, "Abriendo archivo...")

    if isinstance(file_obj, str):
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)

    # Buscar hoja: "Hoja1" o la primera disponible
    if 'Hoja1' in wb.sheetnames:
        ws = wb['Hoja1']
    else:
        ws = wb.worksheets[0]

    if progress_callback:
        progress_callback(0.10, "Detectando formato...")

    # Leer primera fila para detectar headers
    first_row = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        first_row = list(row)
        break

    has_headers = detect_headers(first_row)

    # Determinar mapeo de columnas
    if has_headers:
        headers = [str(c).strip() if c else "" for c in first_row]
        header_map = {h: i for i, h in enumerate(headers)}
        col_idx = {
            'STR': header_map.get('STR', 0),
            'DCS': header_map.get('DCS', 1),
            'VC': header_map.get('VC', 2),
            'TIENDAS': header_map.get('TIENDAS', 3),
            'LINEA': header_map.get('LINEA', 4),
            'SUBLINEA': header_map.get('SUBLINEA', 5),
            'MARCA': header_map.get('MARCA', 6),
            'DESC1': header_map.get('DESC1', 7),
            'ATTR': header_map.get('ATTR', 8),
            'SIZE': header_map.get('SIZE', 9),
            'INVEN': header_map.get('INVEN', 10),
            'VTAS_15': header_map.get('VTAS 15 DIAS', 11),
            'PRECIO': header_map.get('P$T$', 12),
        }
        start_row = 2
    else:
        col_idx = COL_MAP_NO_HEADERS
        start_row = 1

    if progress_callback:
        progress_callback(0.15, "Leyendo registros...")

    # Leer todos los registros
    all_records = []
    row_count = 0
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        vals = list(row)
        if len(vals) < 13:
            continue
        all_records.append({
            'STR': safe_int(vals[col_idx['STR']]),
            'SUBLINEA': safe_str(vals[col_idx['SUBLINEA']]),
            'MARCA': safe_str(vals[col_idx['MARCA']]),
            'DESC1': safe_str(vals[col_idx['DESC1']]),
            'ATTR': safe_str(vals[col_idx['ATTR']]),
            'SIZE': safe_int(vals[col_idx['SIZE']]),
            'INVEN': safe_int(vals[col_idx['INVEN']]),
            'VTAS_15': safe_int(vals[col_idx['VTAS_15']]),
            'PRECIO': safe_float(vals[col_idx['PRECIO']]),
        })
        row_count += 1
        if progress_callback and row_count % 10000 == 0:
            progress_callback(0.15 + 0.55 * min(row_count / 150000, 1.0),
                            f"Leyendo... {row_count:,} registros")

    wb.close()

    if progress_callback:
        progress_callback(0.75, "Detectando línea y tallas...")

    # Detectar línea y tallas
    linea, tallas = detect_line_and_sizes(all_records)

    if progress_callback:
        progress_callback(0.80, "Carga completa")

    return {
        'records': all_records,
        'has_headers': has_headers,
        'linea': linea,
        'tallas': tallas,
        'total_rows': len(all_records),
    }
