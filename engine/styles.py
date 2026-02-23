"""
Constantes de estilos Excel y utilidades compartidas.
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Fills
AZUL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
VERDE_OK = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
AMARILLO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GRIS_CLARO = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
AZUL_CLARO = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
AZUL_TOTAL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

# Fonts
FONT_HEADER = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_SUBTITULO = Font(name="Calibri", bold=True, size=11, color="1F4E79")
FONT_TITULO = Font(name="Calibri", bold=True, size=16, color="1F4E79")
FONT_TITULO_PICKING = Font(name="Calibri", bold=True, size=14, color="1F4E79")
FONT_NORMAL = Font(name="Calibri", size=10)
FONT_BOLD = Font(name="Calibri", bold=True, size=10)
FONT_URGENTE = Font(name="Calibri", bold=True, size=10, color="9C0006")
FONT_ALTA = Font(name="Calibri", bold=True, size=10, color="9C6500")

# Alignment
ALIGN_CENTER = Alignment(horizontal='center')
ALIGN_RIGHT = Alignment(horizontal='right')

# Border
BORDER_THIN = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Column mapping for no-header files (DAMA format)
COL_MAP_NO_HEADERS = {
    'STR': 0, 'DCS': 1, 'VC': 2, 'TIENDAS': 3, 'LINEA': 4,
    'SUBLINEA': 5, 'MARCA': 6, 'DESC1': 7, 'ATTR': 8, 'SIZE': 9,
    'INVEN': 10, 'VTAS_15': 11, 'PRECIO': 12,
}

# Priority order
ORDEN_PRIORIDAD = {'URGENTE': 0, 'ALTA': 1, 'NORMAL': 2, 'BAJA': 3, 'OPORTUNIDAD': 4}


def safe_int(v, d=0):
    if v is None:
        return d
    try:
        return int(v)
    except (ValueError, TypeError):
        return d


def safe_float(v, d=0.0):
    if v is None:
        return d
    try:
        return float(v)
    except (ValueError, TypeError):
        return d


def safe_str(v, d=''):
    if v is None:
        return d
    return str(v).strip()
