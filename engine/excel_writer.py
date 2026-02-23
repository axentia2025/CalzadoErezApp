"""
Generador de Excel con distribución: RESUMEN + DISTRIBUCIÓN COMPLETA + hojas PICKING por tienda.
"""
import openpyxl
from io import BytesIO
from collections import defaultdict
from .styles import (
    AZUL_HEADER, VERDE_OK, AMARILLO, ROJO, GRIS_CLARO, AZUL_CLARO, AZUL_TOTAL,
    FONT_HEADER, FONT_SUBTITULO, FONT_TITULO, FONT_TITULO_PICKING,
    FONT_NORMAL, FONT_BOLD, FONT_URGENTE, FONT_ALTA,
    ALIGN_CENTER, ALIGN_RIGHT, BORDER_THIN,
)


def generate_excel(result, linea, fecha, tallas):
    """Genera Excel completo con distribución.

    Args:
        result: dict retornado por run_distribution()
        linea: "DAMA" o "CABALLERO"
        fecha: string con fecha
        tallas: lista de tallas [23,24,25,26] o [26,27,28,29]

    Returns:
        BytesIO con el archivo Excel listo para descarga
    """
    distribuciones = result['distribuciones']
    resumen_producto = result['resumen_producto']
    summary = result['summary']

    resumen_tienda = summary['resumen_tienda']
    resumen_prioridad = summary['resumen_prioridad']
    total_pares = summary['total_pares']
    total_productos = summary['total_productos']
    total_valor = summary['total_valor']
    total_almacen = summary['total_almacen']
    productos_almacen = summary['productos_almacen']
    pct_distribuido = summary['pct_distribuido']
    total_restante = summary['total_restante']

    wb = openpyxl.Workbook()

    # --- HOJA 1: RESUMEN ---
    ws_res = wb.active
    ws_res.title = "RESUMEN"
    ws_res.sheet_properties.tabColor = "1F4E79"

    ws_res.merge_cells('A1:I1')
    ws_res['A1'] = f"DISTRIBUCIÓN DESDE ALMACÉN - {linea}"
    ws_res['A1'].font = FONT_TITULO
    ws_res.merge_cells('A2:I2')
    ws_res['A2'] = (f"Fecha: {fecha} | Almacén: {total_almacen:,} pares → "
                    f"Distribuido: {total_pares:,} ({pct_distribuido:.0f}%) | "
                    f"Restante: {total_restante:,}")
    ws_res['A2'].font = FONT_SUBTITULO
    ws_res.merge_cells('A3:I3')
    ws_res['A3'] = (f"{productos_almacen} productos | {len(resumen_tienda)} tiendas | "
                    f"Valor: ${total_valor:,.0f} MXN")
    ws_res['A3'].font = FONT_SUBTITULO

    row = 5
    h_res = ['TIENDA', 'PARES', 'PRODUCTOS', 'URGENTE', 'ALTA', 'NORMAL',
             'BAJA', 'OPORTUNIDAD', 'VALOR $']
    for c, h in enumerate(h_res, 1):
        cell = ws_res.cell(row=row, column=c, value=h)
        cell.font = FONT_HEADER
        cell.fill = AZUL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN

    for t in sorted(resumen_tienda.keys()):
        row += 1
        r = resumen_tienda[t]
        vals = [f"Tienda {t}", r['pares'], len(r['productos']),
                r['urgente'], r['alta'], r['normal'], r['baja'],
                r['oportunidad'], r['valor']]
        for c, v in enumerate(vals, 1):
            cell = ws_res.cell(row=row, column=c, value=v)
            cell.font = FONT_NORMAL
            cell.border = BORDER_THIN
            if c >= 2:
                cell.alignment = ALIGN_CENTER
                cell.number_format = '#,##0' if c < 9 else '$#,##0'
        if (row - 5) % 2 == 0:
            for c in range(1, 10):
                ws_res.cell(row=row, column=c).fill = GRIS_CLARO

    # Total
    row += 1
    totals = ['TOTAL', total_pares, total_productos,
              resumen_prioridad.get('URGENTE', 0), resumen_prioridad.get('ALTA', 0),
              resumen_prioridad.get('NORMAL', 0), resumen_prioridad.get('BAJA', 0),
              resumen_prioridad.get('OPORTUNIDAD', 0), total_valor]
    for c, v in enumerate(totals, 1):
        cell = ws_res.cell(row=row, column=c, value=v)
        cell.font = FONT_BOLD
        cell.border = BORDER_THIN
        cell.fill = AZUL_TOTAL
        if c >= 2:
            cell.alignment = ALIGN_CENTER
            cell.number_format = '#,##0' if c < 9 else '$#,##0'

    # Estado del almacén
    row += 3
    ws_res.cell(row=row, column=1, value="ESTADO DEL ALMACÉN").font = FONT_SUBTITULO
    row += 1
    for c, h in enumerate(['PRODUCTO', 'INVENTARIO', 'DISTRIBUIDO', 'RESTANTE', '% DIST'], 1):
        cell = ws_res.cell(row=row, column=c, value=h)
        cell.font = FONT_HEADER
        cell.fill = AZUL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN

    for prod_key in sorted(resumen_producto.keys(),
                           key=lambda k: -resumen_producto[k]['distribuido']):
        row += 1
        info = resumen_producto[prod_key]
        marca, modelo, color = prod_key
        pct = info['distribuido'] / info['inv_original'] * 100 if info['inv_original'] > 0 else 0
        vals = [f"{marca} {modelo} {color}", info['inv_original'],
                info['distribuido'], info['restante'], round(pct, 1)]
        for c, v in enumerate(vals, 1):
            cell = ws_res.cell(row=row, column=c, value=v)
            cell.font = FONT_NORMAL
            cell.border = BORDER_THIN
            if c >= 2:
                cell.alignment = ALIGN_CENTER
                cell.number_format = '#,##0' if c < 5 else '0.0"%"'

    for cl, w in [('A', 20), ('B', 12), ('C', 12), ('D', 14), ('E', 10),
                  ('F', 10), ('G', 10), ('H', 14), ('I', 15)]:
        ws_res.column_dimensions[cl].width = w

    # --- HOJA 2: DISTRIBUCIÓN COMPLETA ---
    ws_dist = wb.create_sheet("DISTRIBUCIÓN COMPLETA")
    ws_dist.sheet_properties.tabColor = "2E75B6"

    d_headers = ['TIENDA', 'SUBLÍNEA', 'MARCA', 'MODELO', 'COLOR', 'TALLA',
                 'QTY ENVIAR', 'PRECIO', 'VALOR', 'PRIORIDAD',
                 'INV ALMACÉN', 'INV TIENDA', 'VTAS 15D']

    for c, h in enumerate(d_headers, 1):
        cell = ws_dist.cell(row=1, column=c, value=h)
        cell.font = FONT_HEADER
        cell.fill = AZUL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN

    for i, d in enumerate(distribuciones, 2):
        vals = [f"Tienda {d['TIENDA']}", d['SUBLINEA'], d['MARCA'], d['MODELO'],
                d['COLOR'], d['TALLA'], d['QTY_ENVIAR'], d['PRECIO'],
                d['QTY_ENVIAR'] * d['PRECIO'], d['PRIORIDAD'],
                d['INV_ALMACEN'], d['INV_TIENDA'], d['VTAS_15D']]
        for c, v in enumerate(vals, 1):
            cell = ws_dist.cell(row=i, column=c, value=v)
            cell.font = FONT_NORMAL
            cell.border = BORDER_THIN
            if c in [6, 7, 11, 12, 13]:
                cell.alignment = ALIGN_CENTER
                cell.number_format = '#,##0'
            elif c in [8, 9]:
                cell.number_format = '$#,##0.00'
                cell.alignment = ALIGN_RIGHT
            elif c == 10:
                cell.alignment = ALIGN_CENTER
                if v == 'URGENTE':
                    cell.fill = ROJO
                    cell.font = FONT_URGENTE
                elif v == 'ALTA':
                    cell.fill = AMARILLO
                    cell.font = FONT_ALTA
                elif v == 'NORMAL':
                    cell.fill = VERDE_OK
                elif v == 'BAJA':
                    cell.fill = AZUL_CLARO

    ws_dist.auto_filter.ref = f"A1:M{len(distribuciones) + 1}"
    for cl, w in [('A', 12), ('B', 14), ('C', 14), ('D', 16), ('E', 14), ('F', 8),
                  ('G', 12), ('H', 12), ('I', 12), ('J', 12), ('K', 12), ('L', 12), ('M', 10)]:
        ws_dist.column_dimensions[cl].width = w
    ws_dist.freeze_panes = 'A2'

    # --- HOJAS POR TIENDA (PICKING LISTS) ---
    for tienda in sorted(resumen_tienda.keys()):
        items = [d for d in distribuciones if d['TIENDA'] == tienda]
        if not items:
            continue

        r = resumen_tienda[tienda]
        n_prods = len(r['productos'])
        ws_t = wb.create_sheet(f"TIENDA {tienda}")
        ws_t.sheet_properties.tabColor = "E74C3C" if r['urgente'] > 0 else "27AE60"

        # Header
        last_col = 5 + len(tallas) + 2  # 5 cols fijas + tallas + total + prioridad + check
        last_col_letter = openpyxl.utils.get_column_letter(last_col + 1)
        ws_t.merge_cells(f'A1:{last_col_letter}1')
        ws_t['A1'] = f"PICKING LIST - TIENDA {tienda} - {linea}"
        ws_t['A1'].font = FONT_TITULO_PICKING

        ws_t.merge_cells(f'A2:{last_col_letter}2')
        ws_t['A2'] = (f"Fecha: {fecha} | {r['pares']:,} pares | "
                      f"{n_prods} productos | Valor: ${r['valor']:,.0f}")
        ws_t['A2'].font = FONT_SUBTITULO

        pick_h = ['#', 'SUBLÍNEA', 'MARCA', 'MODELO', 'COLOR']
        for t in tallas:
            pick_h.append(str(t))
        pick_h.extend(['TOTAL', 'PRIORIDAD', 'CHECK'])

        row = 4
        for c, h in enumerate(pick_h, 1):
            cell = ws_t.cell(row=row, column=c, value=h)
            cell.font = FONT_HEADER
            cell.fill = AZUL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_THIN

        # Agrupar por producto
        productos_tienda = defaultdict(lambda: defaultdict(int))
        producto_info = {}
        for item in items:
            pk = (item['MARCA'], item['MODELO'], item['COLOR'])
            productos_tienda[pk][item['TALLA']] = item['QTY_ENVIAR']
            producto_info[pk] = {'sublinea': item['SUBLINEA'], 'prioridad': item['PRIORIDAD']}

        productos_sorted = sorted(productos_tienda.keys(),
                                  key=lambda k: (producto_info[k]['sublinea'], k[0], k[1], k[2]))

        current_sublinea = None
        num = 0
        for pk in productos_sorted:
            tallas_qty = productos_tienda[pk]
            info = producto_info[pk]

            if info['sublinea'] != current_sublinea:
                if current_sublinea is not None:
                    row += 1
                current_sublinea = info['sublinea']

            row += 1
            num += 1
            total_prod = sum(tallas_qty.values())

            vals = [num, info['sublinea'], pk[0], pk[1], pk[2]]
            for t in tallas:
                vals.append(tallas_qty.get(t, ''))
            vals.extend([total_prod, info['prioridad'], ''])

            for c, v in enumerate(vals, 1):
                cell = ws_t.cell(row=row, column=c, value=v)
                cell.font = FONT_NORMAL
                cell.border = BORDER_THIN
                if 6 <= c <= 5 + len(tallas) + 1:
                    cell.alignment = ALIGN_CENTER
                    if v != '':
                        cell.number_format = '#,##0'
                elif c == 5 + len(tallas) + 2:  # Prioridad
                    cell.alignment = ALIGN_CENTER
                    if v == 'URGENTE':
                        cell.fill = ROJO
                        cell.font = FONT_URGENTE
                    elif v == 'ALTA':
                        cell.fill = AMARILLO
                        cell.font = FONT_ALTA
                    elif v == 'NORMAL':
                        cell.fill = VERDE_OK
                    elif v == 'BAJA':
                        cell.fill = AZUL_CLARO
                elif c == len(vals):
                    cell.alignment = ALIGN_CENTER

        # Total
        row += 2
        total_col = 5 + len(tallas) + 1
        ws_t.cell(row=row, column=total_col - 1, value="TOTAL:").font = FONT_BOLD
        cell_total = ws_t.cell(row=row, column=total_col, value=r['pares'])
        cell_total.font = FONT_BOLD
        cell_total.number_format = '#,##0'
        cell_total.alignment = ALIGN_CENTER
        cell_total.border = BORDER_THIN

        # Firma
        row += 3
        ws_t.merge_cells(f'A{row}:E{row}')
        ws_t.cell(row=row, column=1, value="Preparado por: ________________").font = FONT_NORMAL
        sign_col = 5 + len(tallas) - 1
        sign_end = 5 + len(tallas) + 3
        sign_end_letter = openpyxl.utils.get_column_letter(min(sign_end, last_col + 1))
        sign_start_letter = openpyxl.utils.get_column_letter(sign_col)
        ws_t.merge_cells(f'{sign_start_letter}{row}:{sign_end_letter}{row}')
        ws_t.cell(row=row, column=sign_col, value="Revisado por: ________________").font = FONT_NORMAL

        # Column widths
        for cl, w in [('A', 5), ('B', 13), ('C', 13), ('D', 15), ('E', 13)]:
            ws_t.column_dimensions[cl].width = w
        for i, t in enumerate(tallas):
            col_letter = openpyxl.utils.get_column_letter(6 + i)
            ws_t.column_dimensions[col_letter].width = 6
        ws_t.column_dimensions[openpyxl.utils.get_column_letter(6 + len(tallas))].width = 8
        ws_t.column_dimensions[openpyxl.utils.get_column_letter(7 + len(tallas))].width = 12
        ws_t.column_dimensions[openpyxl.utils.get_column_letter(8 + len(tallas))].width = 8

        ws_t.page_setup.orientation = 'landscape'
        ws_t.page_setup.fitToWidth = 1
        ws_t.page_setup.fitToHeight = 0
        ws_t.sheet_properties.pageSetUpPr.fitToPage = True

    # Guardar a BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
